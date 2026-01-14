from openpyxl import load_workbook
from io import BytesIO
import re

CELL_REF = re.compile(r'\b[A-Z]{1,3}[0-9]{1,7}\b')

def extract_dependencies(formula):
    if not formula:
        return []
    return list(set(CELL_REF.findall(formula)))

def parse_excel(file):
    
    print("ZIP HEADER:", file[:4])
    file_formula = BytesIO(file)
    file_value = BytesIO(file)

    wb_formula = load_workbook(file_formula, data_only=False)
    wb_value = load_workbook(file_value, data_only=True)

    result = {}

    for sheet_f, sheet_v in zip(wb_formula.worksheets, wb_value.worksheets):
        
        sheet_name = sheet_f.title
        result[sheet_name] = {}

        for row in sheet_f.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                # Extract formula (if the cell has a formula)
                formula = cell.value if cell.data_type == "f" else None
                
                # Get the corresponding computed value from the values sheet
                value_cell = sheet_v[cell.coordinate]
                value = value_cell.value
                
                # Get the data type of the *computed* value
                # (usually more useful than the formula cell's data type)
                data_type = value_cell.data_type if value_cell else None
                
                # Optional: you can also get the formula cell's data type if needed
                formula_data_type = cell.data_type

                result[sheet_name][cell.coordinate] = {
                    "value": value,
                    "formula": formula,
                    "data_type": data_type,           # ← added: type of the computed value
                    "formula_data_type": formula_data_type,  # optional
                    "dependencies": extract_dependencies(formula)
                }

    return result
